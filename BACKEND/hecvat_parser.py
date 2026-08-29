import os
import re

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import config


# maps a REQU answer to whatever it should switch off elsewhere in the workbook; mirrors the Auto Responses sheet logic
ROUTING_RULES = {
    ("REQU-01", "No"):  ["sheet:Product", "sheet:Infrastructure"],
    ("REQU-02", "No"):  ["sheet:IT Accessibility"],
    ("REQU-03", "No"):  ["prefix:CONS"],
    ("REQU-04", "No"):  ["sheet:AI"],
    ("REQU-05", "No"):  ["prefix:HIPA"],
    ("REQU-06", "No"):  ["prefix:PCID"],
    ("REQU-07", "No"):  ["prefix:ONPR"],
    ("REQU-08", "No"):  ["sheet:Privacy"],
}

ROUTING_RULE_SCOPES = {
    "REQU-01": "cloud",
    "REQU-02": "accessibility",
    "REQU-04": "ai",
    "REQU-05": "hipaa",
    "REQU-06": "pci_dss",
    "REQU-07": "on_premises",
    "REQU-08": "privacy",
}

ANALYST_SHEETS = {
    "Institution Evaluation",
    "High-Risk Evaluation",
    "Privacy Analyst Evaluation",
    "Analyst Reference",
    "Questions",
    "Auto Responses",
    "(backend scoring)",
    "START HERE",
}

METADATA_PREFIXES = {"GNRL", "COMP", "REQU", "AIQU"}


def _norm(s) -> str:
    # squash to lowercase alphanumerics so sheet/column names still match despite spacing, punctuation, or casing drift
    return re.sub(r"[^a-z0-9]", "", str(s if s is not None else "").lower())


def _prefix(control_id: str) -> str:
    return control_id.split("-")[0].upper() if "-" in control_id else control_id[:4].upper()


def _find_sheet(wb, candidates: list[str]):
    # resolve a sheet by normalized name — exact first, then either-way substring, so "(backend scoring)" ~ "Backend Scoring"
    norm_to_real = {}
    for real in wb.sheetnames:
        norm_to_real.setdefault(_norm(real), real)
    cands = [_norm(c) for c in candidates if _norm(c)]
    for c in cands:
        if c in norm_to_real:
            return wb[norm_to_real[c]]
    for c in cands:
        for nkey, real in norm_to_real.items():
            if c and (c in nkey or nkey in c):
                return wb[real]
    return None


def _resolve_columns(rows, alias_map: dict, scan: int = 25):
    # find the header row (the one resolving the most fields) and map each logical field to a column index by its label
    best_idx, best_cols = None, {}
    for idx in range(min(scan, len(rows))):
        row = rows[idx] or ()
        cells = [(_norm(c), j) for j, c in enumerate(row)]
        cols, used = {}, set()
        # exact matches first so a precise header claims the column before any fuzzy fallback can grab it
        for mode in ("exact", "fuzzy"):
            for field, aliases in alias_map.items():
                if field in cols:
                    continue
                for a in aliases:
                    na = _norm(a)
                    if not na:
                        continue
                    # a 2-char alias like "id" would fuzzy-match "guidance" — only let short tokens match exactly
                    if mode == "fuzzy" and len(na) < 3:
                        continue
                    for nc, j in cells:
                        if j in used or not nc:
                            continue
                        hit = (nc == na) if mode == "exact" else (nc.startswith(na) or na in nc)
                        if hit:
                            cols[field] = j
                            used.add(j)
                            break
                    if field in cols:
                        break
        if len(cols) > len(best_cols):
            best_idx, best_cols = idx, cols
    return best_idx, best_cols


# section-membership flags are their own columns in the Questions sheet; matched exactly since they're short tokens
_SECTION_FLAG_ALIASES = {
    "in_start":   ["Start"],
    "in_org":     ["Org", "Organization"],
    "in_product": ["Product"],
    "in_infra":   ["Infra", "Infrastructure"],
    "in_access":  ["Access", "Accessibility"],
    "in_case":    ["Case", "Case-Specific"],
    "in_ai":      ["AI"],
    "in_privacy": ["Privacy"],
}


def _resolve_section_flags(header_row) -> dict:
    flags = {}
    norm_cells = {}
    for j, c in enumerate(header_row or ()):
        nc = _norm(c)
        if nc and nc not in norm_cells:
            norm_cells[nc] = j
    for field, aliases in _SECTION_FLAG_ALIASES.items():
        for a in aliases:
            if _norm(a) in norm_cells:
                flags[field] = norm_cells[_norm(a)]
                break
    return flags


def _flag_truthy(v) -> bool:
    # excel stores membership as 1.0 (or blank), so anything zero/blank/None/false means "not in this sheet"
    return v not in (None, "", 0, 0.0, "0", "0.0", "False", "false", "FALSE")


def _is_high_risk(v) -> bool:
    # High Risk arrives as a real bool, but survive a stringified "True"/"Yes"/"1" too
    if isinstance(v, bool):
        return v
    return str(v).strip().lower() in ("true", "yes", "y", "1", "1.0", "high")


_BACKEND_ALIASES = {
    "control_id":    ["New ID", "ID", "Question ID", "Question #"],
    "answer":        ["Vendor Response", "Response", "Answer"],
    "score_mapping": ["Score Mapping"],
}

_QUESTIONS_ALIASES = {
    "control_id":         ["New ID", "ID", "Question ID", "Question #"],
    "question":           ["Question", "Question Text"],
    "score_mapping":      ["Score Mapping"],
    "score_location":     ["Score Location"],
    "guidance":           ["Standard Guidance", "Guidance"],
    "compliant_response": ["Compliant Response"],
    "default_importance": ["Default Importance", "Importance"],
    "default_weight":     ["Default Weight", "Weight"],
}

# the official EDUCAUSE HECVAT Full/Lite keeps everything on one flat sheet — vendor answer sits in V_Answer, no separate scoring sheet
_QUESTIONS_FLAT_ALIASES = {
    "control_id":         ["ID", "Question ID", "New ID", "Question #"],
    "question":           ["Question", "Question Text"],
    "answer":             ["V_Answer", "Vendor Response", "Vendor Answer", "Response", "Answer"],
    "compliant_response": ["C_Answer", "Compliant Answer", "Compliant Response"],
    "guidance":           ["Additional Info", "Additional Information", "Standard Guidance", "Guidance"],
    "category":           ["Category", "Section"],
    "high_risk":          ["High Risk", "High-Risk"],
    "weight":             ["Weight", "Default Weight"],
}


def _build_disabled(routing_answers: dict) -> tuple[set, set, list[dict]]:
    disabled_sheets   = set()
    disabled_prefixes = set()
    events = []

    # case-insensitive compare since vendors can type "no"/"No"/"NO" and mean the same thing
    for (question_id, trigger_answer), rules in ROUTING_RULES.items():
        vendor_answer = routing_answers.get(question_id, "").strip().lower()
        if vendor_answer == trigger_answer.lower():
            events.append({
                "question_id": question_id,
                "answer": routing_answers.get(question_id, ""),
                "rules": list(rules),
                "scope": ROUTING_RULE_SCOPES.get(question_id),
            })
            for rule in rules:
                if rule.startswith("sheet:"):
                    disabled_sheets.add(rule[6:])
                elif rule.startswith("prefix:"):
                    disabled_prefixes.add(rule[7:].upper())

    return disabled_sheets, disabled_prefixes, events


def _control_scopes(control_id: str, sheet: str, question: str) -> set[str]:
    prefix = _prefix(control_id)
    text = f"{sheet} {question}".lower()
    scopes = set()
    if prefix == "HIPA":
        scopes.add("hipaa")
    if prefix == "PCID":
        scopes.add("pci_dss")
    if prefix == "ONPR":
        scopes.add("on_premises")
    if prefix.startswith("AI") or sheet == "AI":
        scopes.add("ai")
    if prefix == "ITAC" or sheet == "IT Accessibility":
        scopes.add("accessibility")
    if prefix in {"PRIV", "PCOM", "PDOC", "PTHP", "PCHG", "PDAT", "PRPO", "DRPV", "INTL", "PRGN"} or sheet == "Privacy":
        scopes.add("privacy")
    if "ferpa" in text:
        scopes.add("ferpa")
    if "coppa" in text:
        scopes.add("coppa")
    if "cloud" in text or sheet in {"Product", "Infrastructure"}:
        scopes.add("cloud")
    return scopes


def _routing_events_for_control(events: list[dict], sheet: str, prefix: str) -> list[dict]:
    matched = []
    for event in events:
        for rule in event["rules"]:
            if rule == f"sheet:{sheet}" or rule == f"prefix:{prefix}":
                matched.append(event)
                break
    return matched


def _read_backend_scoring(ws) -> dict:
    # the scoring sheet holds the vendor's actual answers (Vendor Response column), keyed by control id
    answers = {}
    if ws is None:
        return answers
    rows = list(ws.iter_rows(values_only=True))
    hdr_idx, cols = _resolve_columns(rows, _BACKEND_ALIASES)
    if hdr_idx is None or "control_id" not in cols or "answer" not in cols:
        print("  [Parser] Warning: could not locate ID/response columns in the scoring sheet")
        return answers
    ci, ai = cols["control_id"], cols["answer"]
    smi = cols.get("score_mapping")
    for source_row, row in enumerate(rows[hdr_idx + 1:], start=hdr_idx + 2):
        try:
            if not row or ci >= len(row) or not row[ci]:
                continue
            cid = str(row[ci]).strip()
            answer = str(row[ai] or "").strip() if ai < len(row) else ""
            score_map = str(row[smi] or "").strip() if smi is not None and smi < len(row) else ""
            answers[cid] = {
                "answer": answer,
                "score_mapping": score_map,
                "source_worksheet": ws.title,
                "source_row": source_row,
                "source_cell": f"{ws.title}!{get_column_letter(ai + 1)}{source_row}",
            }
        except Exception as e:
            print(f"  [Parser] Warning: skipping bad scoring row: {e}")
    return answers


def _read_questions_master(ws) -> dict:
    controls = {}
    if ws is None:
        return controls
    rows = list(ws.iter_rows(values_only=True))
    hdr_idx, cols = _resolve_columns(rows, _QUESTIONS_ALIASES)
    if hdr_idx is None or "control_id" not in cols or "question" not in cols:
        print("  [Parser] Warning: could not locate ID/Question columns in the Questions sheet")
        return controls
    flags = _resolve_section_flags(rows[hdr_idx])
    ci = cols["control_id"]
    for source_row, row in enumerate(rows[hdr_idx + 1:], start=hdr_idx + 2):
        try:
            if not row or ci >= len(row) or not row[ci]:
                continue
            control_id = str(row[ci]).strip()
            if not control_id or control_id.lower().startswith("end"):
                continue

            def cell(field, default=""):
                j = cols.get(field)
                return str(row[j] or default).strip() if j is not None and j < len(row) else default

            entry = {
                "control_id":        control_id,
                "question":          cell("question"),
                "score_mapping":     cell("score_mapping"),
                "score_location":    cell("score_location"),
                "guidance":          cell("guidance"),
                "compliant_response":cell("compliant_response"),
                "default_importance":cell("default_importance"),
                "default_weight":    cell("default_weight", "10"),
                "requirement_worksheet": ws.title,
                "requirement_row":       source_row,
                "requirement_cell":      f"{ws.title}!{get_column_letter(cols['question'] + 1)}{source_row}",
            }
            # every in_* key must exist for _infer_sheet/_all_sheets; a flag column we couldn't find defaults to False
            for fk in ("in_start", "in_org", "in_product", "in_infra", "in_access", "in_case", "in_ai", "in_privacy"):
                j = flags.get(fk)
                entry[fk] = _flag_truthy(row[j]) if (j is not None and j < len(row)) else False
            controls[control_id] = entry
        except Exception as e:
            print(f"  [Parser] Warning: skipping bad Questions row: {e}")
    return controls


def _read_questions_flat(ws) -> list[dict]:
    # single-sheet HECVAT: build finished control dicts straight from one row each, no routing/scoring cross-references
    controls = []
    if ws is None:
        return controls
    rows = list(ws.iter_rows(values_only=True))
    hdr_idx, cols = _resolve_columns(rows, _QUESTIONS_FLAT_ALIASES)
    # this layout only makes sense if the vendor's answer column is actually present
    if hdr_idx is None or "control_id" not in cols or "question" not in cols or "answer" not in cols:
        return controls
    ci = cols["control_id"]
    for source_row, row in enumerate(rows[hdr_idx + 1:], start=hdr_idx + 2):
        try:
            if not row or ci >= len(row) or not row[ci]:
                continue
            control_id = str(row[ci]).strip()
            if not control_id or control_id.lower().startswith("end"):
                continue

            def cell(field, default=""):
                j = cols.get(field)
                return str(row[j] or default).strip() if j is not None and j < len(row) else default

            question = cell("question")
            if not question:
                continue

            prefix = _prefix(control_id) if "-" in control_id else ""
            # GNRL/COMP/REQU/AIQU are metadata/scoping rows, not assessable controls
            if prefix in METADATA_PREFIXES:
                continue

            # flat layout has no routing, so a vendor "N/A" is a genuine response — keep it, don't relabel as auto-excluded
            answer = cell("answer") or "Not answered"
            section = cell("category") or _get_section(control_id)
            scopes = _control_scopes(control_id, section, question)
            configured = {
                scope: config.INSTITUTIONAL_APPLICABILITY.get(scope, True)
                for scope in sorted(scopes)
            }
            institutionally_excluded = bool(scopes) and all(
                not applicable for applicable in configured.values()
            )

            hr_j = cols.get("high_risk")
            hr_raw = row[hr_j] if (hr_j is not None and hr_j < len(row)) else None

            controls.append({
                "control_id":         control_id,
                "section":            section,
                "sheet":              section,
                "question":           question,
                "answer":             answer,
                "additional_info":    "",
                "analyst_notes":      "",
                "guidance":           cell("guidance"),
                # High Risk flag drives criticality here; trailing * is the older convention
                "is_critical":        _is_high_risk(hr_raw) or question.rstrip().endswith("*"),
                "weight":             _parse_weight(cell("weight", "10")),
                "score_mapping":      "",
                "compliant_response": cell("compliant_response"),
                "is_na_routed":       False,
                "source_worksheet":   ws.title,
                "source_row":         source_row,
                "source_cell":        f"{ws.title}!{get_column_letter(cols['answer'] + 1)}{source_row}",
                "requirement_cell":   f"{ws.title}!{get_column_letter(cols['question'] + 1)}{source_row}",
                "institutional_scopes": configured,
                "institutional_applicability": not institutionally_excluded,
                "applicability_status": "EXCLUDED" if institutionally_excluded else "APPLICABLE",
                "applicability_audit": {
                    "vendor_answer": answer,
                    "institutional_configuration": configured,
                    "routing_rules": [],
                    "affected_control": control_id,
                    "reason": (
                        "All configured institutional scopes for this control are disabled."
                        if institutionally_excluded
                        else "Institutional scope applies."
                    ),
                    "audit_event": "institutional_applicability_evaluated",
                },
            })
        except Exception as e:
            print(f"  [Parser] Warning: skipping bad Questions row: {e}")
    return controls


def _infer_sheet(control: dict) -> str:
    # order matters here — a control can have multiple in_* flags set, so the most specific sheet wins first
    if control["in_privacy"]: return "Privacy"
    if control["in_ai"]:      return "AI"
    if control["in_case"]:    return "Case-Specific"
    if control["in_access"]:  return "IT Accessibility"
    if control["in_infra"]:   return "Infrastructure"
    if control["in_product"]: return "Product"
    if control["in_org"]:     return "Organization"
    if control["in_start"]:   return "START HERE"
    return "Unknown"


def _all_sheets(control: dict) -> set:
    # returns the full set of sheets this control belongs to, for disable/enable decisions (H4)
    mapping = [
        ("in_privacy", "Privacy"),
        ("in_ai",      "AI"),
        ("in_case",    "Case-Specific"),
        ("in_access",  "IT Accessibility"),
        ("in_infra",   "Infrastructure"),
        ("in_product", "Product"),
        ("in_org",     "Organization"),
        ("in_start",   "START HERE"),
    ]
    return {sheet for key, sheet in mapping if control.get(key)}


PREFIX_TO_SECTION = {
    "DOCU": "Documentation",
    "THRD": "Third-Party Management",
    "CHNG": "Change Management",
    "AAAI": "Access Control",
    "DATA": "Data Security",
    "DCTR": "Data Center",
    "MOBL": "Mobile",
    "NETW": "Network Security",
    "VULN": "Vulnerability Management",
    "INCD": "Incident Response",
    "BKUP": "Backup & Recovery",
    "ENCR": "Encryption",
    "SOFT": "Software Development",
    "PHYS": "Physical Security",
    "PRIV": "Privacy",
    "CONS": "Consulting",
    "HIPA": "HIPAA",
    "PCID": "PCI-DSS",
    "ONPR": "On-Premises",
    "AIPL": "AI Policy",
    "AIFN": "AI Functionality",
    "AISC": "AI Security",
    "AIPV": "AI Privacy",
    "ITAC": "IT Accessibility",
    "HFIH": "Incident Response",
    "FIDP": "Firewall & IDS/IPS",
    "PPPR": "Policies, Processes & Procedures",
    "OPEM": "Operational & Emerging Tech",
    "APPL": "Application Security",
    "AIGN": "AI Governance",
    "AIML": "AI/ML Data Separation",
    "AILM": "LLM Privileges",
    "DPAI": "AI & Data Privacy",
    "PCOM": "Privacy Compliance",
    "PDOC": "Privacy Documentation",
    "PTHP": "Privacy Third Parties",
    "PCHG": "Privacy Change Management",
    "PDAT": "Personal Data Processing",
    "PRPO": "Privacy Risk & Programme",
    "DRPV": "Data Privacy Impact Assessment",
    "INTL": "International",
    "PRGN": "FERPA/COPPA",
}

def _get_section(control_id: str) -> str:
    return PREFIX_TO_SECTION.get(_prefix(control_id), _prefix(control_id))


def parse_hecvat_excel(excel_path: str, debug: bool = False, require_answers: bool = True) -> list[dict]:
    wb = load_workbook(excel_path, read_only=True, data_only=True)
    try:
        # fuzzy sheet names to survive version drift; bare "scoring" would match "Scoring Guide"
        scoring_ws   = _find_sheet(wb, ["(backend scoring)", "backend scoring"])
        questions_ws = _find_sheet(wb, ["Questions", "Questions Master", "Question Master"])

        backend_answers = _read_backend_scoring(scoring_ws)

        # no usable backend-scoring sheet (missing, or a stray tab with no ID/answer cols) → flat EDUCAUSE Full/Lite layout
        if not backend_answers:
            flat = _read_questions_flat(questions_ws)
            if flat:
                if require_answers:
                    assessable = [c for c in flat if c.get("institutional_applicability", True)]
                    answered = sum(1 for c in assessable if c["answer"] not in ("", "Not answered"))
                    blank_ratio = 1.0 - (answered / len(assessable)) if assessable else 0.0
                    if blank_ratio >= 0.6:
                        raise ValueError(
                            "this HECVAT has almost no vendor answers — nearly every response cell is blank. "
                            "upload a completed one."
                        )
                print(f"  [Parser] Flat HECVAT layout — assessable controls: {len(flat)}")
                for control in flat:
                    control["source_workbook"] = os.path.basename(excel_path)
                return flat
            # neither a scoring sheet nor a readable flat Questions sheet — genuinely can't find vendor answers
            raise ValueError(
                "no vendor responses found — expected either a '(backend scoring)' sheet (HECVAT 4.x) "
                "or a flat 'Questions' sheet with a vendor-answer column (EDUCAUSE HECVAT Full/Lite). "
                "This workbook has neither."
            )

        # routing questions (REQU/AIQU) sit in that same scoring sheet — pull their answers to know which sections to disable
        routing_answers = {cid: v["answer"] for cid, v in backend_answers.items() if _prefix(cid) in ("REQU", "AIQU")}
        if debug:
            print(f"  [Parser] Routing answers: {routing_answers}")

        disabled_sheets, disabled_prefixes, routing_events = _build_disabled(routing_answers)
        if debug:
            print(f"  [Parser] Disabled sheets: {disabled_sheets}")
            print(f"  [Parser] Disabled prefixes: {disabled_prefixes}")

        questions = _read_questions_master(questions_ws)
        # H5: a missing/renamed/unreadable Questions sheet yields an empty master — fail loud instead of assessing 0 controls
        if not questions:
            raise ValueError(
                "Questions master sheet missing or unreadable — cannot read any controls. "
                "Check the workbook has a 'Questions' sheet in the HECVAT layout."
            )
        if debug:
            print(f"  [Parser] Questions loaded: {len(questions)}  Backend answers: {len(backend_answers)}")

        controls = []
        skipped_metadata    = 0
        skipped_disabled    = 0
        skipped_notscored   = 0
        skipped_noquestion  = 0
        skipped_naonly      = 0  # controls excluded by routing auto-fill (tracked separately from disabled)

        for control_id, q in questions.items():
            if not q["question"]:
                skipped_noquestion += 1
                continue

            prefix = _prefix(control_id) if "-" in control_id else ""
            if prefix in METADATA_PREFIXES:
                skipped_metadata += 1
                continue

            # normalize so "N/A" (with slash) and "Not Scored" spacing variants still match
            score_loc = _norm(q["score_location"])
            score_map = _norm(q["score_mapping"])
            if score_loc == "notscored" or score_map in ("na", "notapplicable"):
                skipped_notscored += 1
                continue

            sheet = _infer_sheet(q)

            if sheet in ANALYST_SHEETS:
                skipped_metadata += 1
                continue

            # H4: only exclude if ALL sheets this control belongs to are disabled
            all_s = _all_sheets(q)
            vendor_routing_match = (
                (bool(all_s) and all_s.issubset(disabled_sheets))
                or (not all_s and sheet in disabled_sheets)
                or prefix in disabled_prefixes
            )
            matched_routing = (
                _routing_events_for_control(routing_events, sheet, prefix)
                if vendor_routing_match else []
            )
            scopes = _control_scopes(control_id, sheet, q["question"])
            configured = {
                scope: config.INSTITUTIONAL_APPLICABILITY.get(scope, True)
                for scope in sorted(scopes)
            }
            institutionally_excluded = bool(scopes) and all(
                not applicable for applicable in configured.values()
            )

            ba = backend_answers.get(control_id, {})
            answer = ba.get("answer", "Not answered") or "Not answered"

            # a vendor may auto-fill N/A — still assessable if the scope applies to us
            is_na = answer.lower() == "n/a" or answer.lower().startswith("based on the response")

            controls.append({
                "control_id":         control_id,
                "section":            _get_section(control_id),
                "sheet":              sheet,
                "question":           q["question"],
                "answer":             "N/A (auto-excluded by routing)" if is_na else answer,
                "additional_info":    "",   # not available from backend scoring
                "analyst_notes":      "",
                "guidance":           q["guidance"],
                # L3: anchor is_critical to a trailing * (not substring) to avoid false positives mid-question
                "is_critical":        q["question"].rstrip().endswith("*") or q["default_importance"] in ("Critical", "High"),
                "weight":             _parse_weight(q["default_weight"]),
                "score_mapping":      q["score_mapping"],
                "compliant_response": q["compliant_response"],
                "is_na_routed":       is_na,
                "source_workbook":    os.path.basename(excel_path),
                "source_worksheet":   ba.get("source_worksheet") or q.get("requirement_worksheet") or sheet,
                "source_row":         ba.get("source_row") or q.get("requirement_row"),
                "source_cell":        ba.get("source_cell") or q.get("requirement_cell"),
                "requirement_cell":   q.get("requirement_cell"),
                "institutional_scopes": configured,
                "institutional_applicability": not institutionally_excluded,
                "applicability_status": "EXCLUDED" if institutionally_excluded else "APPLICABLE",
                "applicability_audit": {
                    "vendor_answer": [event["answer"] for event in matched_routing],
                    "institutional_configuration": configured,
                    "routing_rules": [
                        rule for event in matched_routing for rule in event["rules"]
                    ],
                    "affected_control": control_id,
                    "reason": (
                        "All configured institutional scopes for this control are disabled."
                        if institutionally_excluded
                        else (
                            "Vendor routing requested N/A, but institutional scope remains applicable."
                            if vendor_routing_match or is_na
                            else "Institutional scope applies."
                        )
                    ),
                    "audit_event": "institutional_applicability_evaluated",
                },
            })
            if institutionally_excluded:
                skipped_disabled += 1

        print(f"  [Parser] Skipped: {skipped_metadata} metadata, "
              f"{skipped_disabled} routing-disabled, "
              f"{skipped_notscored} not-scored, "
              f"{skipped_noquestion} no-question")
        print(f"  [Parser] Assessable controls: {len(controls)}")

        # an uncalculated workbook reads back all blank — fail loudly rather than assess nothing
        if controls and require_answers:
            assessable = [c for c in controls if c.get("institutional_applicability", True)]
            answered = sum(1 for c in assessable if c["answer"] not in ("", "Not answered", "N/A (auto-excluded by routing)"))
            blank_ratio = 1.0 - (answered / len(assessable)) if assessable else 0.0
            if blank_ratio >= 0.6:
                raise ValueError(
                    "workbook appears uncalculated — over 60% of answer cells are blank. "
                    "Open and save the file in Excel or LibreOffice so the formulas are evaluated, "
                    "then re-upload. Otherwise, vendor answers cannot be read."
                )

        return controls
    finally:
        wb.close()


def _parse_weight(w) -> int:
    try:
        return int(float(str(w)))
    except Exception:
        return 10


def hecvat_control_to_text(c: dict) -> str:
    parts = [
        f"Control: {c['control_id']}",
        f"Section: {c['section']}",
        f"Question: {c['question']}",
    ]
    if c.get("guidance"):
        parts.append(f"Guidance: {c['guidance']}")
    if c.get("compliant_response"):
        parts.append(f"Compliant Response: {c['compliant_response']}")
    return "\n".join(parts)


def hecvat_control_to_metadata(c: dict) -> dict:
    return {
        "control_id": c["control_id"],
        "section":    c["section"],
        "sheet":      c.get("sheet", ""),
        "source":     "hecvat_template",
    }
